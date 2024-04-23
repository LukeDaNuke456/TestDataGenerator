import random

from openpyxl import load_workbook

# import workbookinitialization

class Vehicle:
    
    def __init__(self, year, make, model,  mileage, price):
        self.make = make
        self.model = model
        self.year = year
        self.mileage = mileage
        self.price = price

    def generate_random_mileage(self):
        if self.year == 2024:
            return random.randint(0, 20000)  # New vehicles may have low mileage
        elif self.year == 2023:
            return random.randint(10000, 40000)  # New vehicles may have low mileage
        elif self.year == 2022:
            return random.randint(30000, 80000)  # New vehicles may have low mileage
        elif self.year == 2021:
            return random.randint(50000, 100000)  # One year old vehicles may have higher mileage
        else:
            return random.randint(100000, 200000)  # Older vehicles may have even higher mileage
        
    def determine_cheverlet_price(self):
        if self.year == 2024 and self.make == 'Cheverlet' and self.model == 'Silverado 1500':
            if self.mileage in range(0,1000): 
                return random.randint(38000, 65000)
            elif self.mileage in range (1001, 10000):
                return random.randint(30000, 50000)
            elif self.mileage in range (10001, 50000):
                return random.randint(25000, 45000)
            elif self.mileage in range (50001, 100000):
                return random.randint(20000, 38000)
            elif self.mileage > 101000:
                return random.randint(12000, 25000)
        if self.year == 2023 and self.make == 'Cheverlet' and self.model == 'Silverado 1500':
            if self.mileage in range(0,1000): 
                return random.randint(38000, 65000)
            elif self.mileage in range (1001, 10000):
                return random.randint(30000, 50000)
            elif self.mileage in range (10001, 50000):
                return random.randint(25000, 45000)
            elif self.mileage in range (50001, 100000):
                return random.randint(20000, 38000)
            elif self.mileage > 101000:
                return random.randint(12000, 25000)
        if self.year == 2022 and self.make == 'Cheverlet' and self.model == 'Silverado 1500':
            if self.mileage in range(0,1000): 
                return random.randint(38000, 65000)
            elif self.mileage in range (1001, 10000):
                return random.randint(30000, 50000)
            elif self.mileage in range (10001, 50000):
                return random.randint(25000, 45000)
            elif self.mileage in range (50001, 100000):
                return random.randint(20000, 38000)
            elif self.mileage > 101000:
                return random.randint(12000, 25000)
            
            
def autoPopulateVehicles(num_vehicles):
    vehicles = {}

    for i in range(1, num_vehicles + 1):

        vehicles[i] = Vehicle(year=0, make='', model='', mileage=0, price=0)

    return vehicles

def setVehicleYear(vehicles):

    vehicleYear = [2021, 2022, 2023, 2024] 

    for vehicle_id, vehicle in vehicles.items():

        random_make_number = random.randint(0,3)

        if vehicle.year == 0:
            
            vehicle.year = vehicleYear[random_make_number]

def setVehicleMake(vehicles):

    vehicleMake = ['Ford', 'Cheverlet', 'Honda', 'Toyota'] 
    
    for vehicle_id, vehicle in vehicles.items():

        random_make_number = random.randint(0,3)

        if vehicle.make == '':
            
            vehicle.make = vehicleMake[random_make_number]
        
        else: 

            vehicle.model = ''

def setVehicleModels(vehicles):
    
    for vehicle_id, vehicle in vehicles.items():

        # print(str(vehicle.make))

        if vehicle.make == 'Honda':

            vehicle.model = setHondaModels()
            
        elif vehicle.make == 'Toyota':

            vehicle.model = setToyotaModels()
        
        elif vehicle.make == 'Cheverlet':

            vehicle.model = setCheverletModels()
        
        elif vehicle.make == 'Ford':

            vehicle.model = setFordModels()
        
def setToyotaModels():

    toyotaModel = ['Camry', 'Tacoma', 'Tundra', 'Rav4']
    random_make_number = random.randint(0, len(toyotaModel) - 1)
    return toyotaModel[random_make_number]

def setFordModels():

    fordModel = ['F-150', 'F-250', 'Ranger', 'Bronco']
    random_make_number = random.randint(0, len(fordModel) - 1)
    return fordModel[random_make_number]

def setHondaModels():

    hondaModel = ['Accord', 'Civic', 'Passport', 'Ridgeline']
    random_make_number = random.randint(0, len(hondaModel) - 1)
    return hondaModel[random_make_number]

def setCheverletModels():

    cheverletModels = ['Silverado 1500', 'Camaro', 'Impala', 'Malibu']
    random_make_number = random.randint(0, len(cheverletModels) - 1)
    return cheverletModels[random_make_number]

def setModelsPricing(vehicles):

    decrease_per_mile = 0.2
    decrease_per_year = 500
    current_year = 2024

    modelsBasePrices = {

        ("Cheverlet", "Silverado 1500"): 65000,
        ("Cheverlet", "Camaro"): 45000,
        ("Cheverlet", "Impala"): 30000, 
        ("Cheverlet", "Malibu"): 27000, 

        ("Honda", "Ridgeline"): 46000,
        ("Honda", "Passport"): 43000,
        ("Honda", "Civic"): 27000, 
        ("Honda", "Accord"): 34000,  

        ("Toyota", "Tundra"): 60000,
        ("Toyota", "Tacoma"): 54000,
        ("Toyota", "Rav4"): 33000, 
        ("Toyota", "Camry"): 30000, 

        ("Ford", "F-150"): 55000,
        ("Ford", "F-250"): 70000,
        ("Ford", "Ranger"): 41000, 
        ("Ford", "Bronco"): 60000        

    }

    for vehicle_id, vehicle in vehicles.items():
        
        if (vehicle.make, vehicle.model) in modelsBasePrices: 

            age = current_year - vehicle.year
            
            basePrice = modelsBasePrices[vehicle.make, vehicle.model]

            adjusted_price = basePrice - (vehicle.mileage * decrease_per_mile)

            adjusted_price -= age * decrease_per_year

            adjusted_price = max(adjusted_price, 0)

            vehicle.price = int(adjusted_price)
        
        else: 
            vehicle.price = 0



    
def setVehicleMiles(vehicles):

    for vehicle_id, vehicle in vehicles.items():

        if vehicle.mileage == 0:
            vehicle.mileage = vehicle.generate_random_mileage()

def writeToWorkbook(vehicles):

    existing_wb = load_workbook("vehicle_data.xlsx")
    ws = existing_wb.active
    header = ["ID", "Year", "Make", "Model", "Mileage", "Price"]
    ws.append(header)

    for vehicle_id, vehicle in vehicles.items():
        ws.append([vehicle_id, vehicle.year, vehicle.make, vehicle.model, vehicle.mileage, vehicle.price])
    
    existing_wb.save("vehicle_data.xlsx")

def main():

    while True: 

        print("How many vehicle's do you want in your inventory? : ")
        num_of_vehicles = input()
            
        if num_of_vehicles.isdigit():

            num_of_vehicles = int(num_of_vehicles)

            vehicles = autoPopulateVehicles(num_of_vehicles)
            setVehicleYear(vehicles)
            setVehicleMake(vehicles)
            setVehicleModels(vehicles)
            setVehicleMiles(vehicles)
            setModelsPricing(vehicles)
            writeToWorkbook(vehicles)

            return vehicles

            # print("---------------------------------\n")

            # for vehicle_id, vehicle in vehicles.items():
            #     print(f"Vehicle ID: {vehicle_id}")
            #     print(f"Year: {vehicle.year}")
            #     print(f"Make: {vehicle.make}")
            #     print(f"Model: {vehicle.model}")
            #     print(f"Mileage: {vehicle.mileage} miles")
            #     print(f"Price: ${vehicle.price} \n")
            
            # print("---------------------------------\n")

            break

        else: 

            print("Not a number. How many vehicle's do you want in your inventory? : ")  

if __name__ == "__main__":
    main()
# def populate_arrays():

   




